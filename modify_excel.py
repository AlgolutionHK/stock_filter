import openpyxl
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
import pandas as pd
import shutil, os, glob


def signal():
    time = datetime.now()
    prev_time = time - timedelta(days=1)
    path = time.strftime("%b_%Y") + ".xlsx"
    if not os.path.exists(path):
        # Get the newest XLSX file
        newest_xlsx_file = max(glob.glob("*.xlsx"), key=os.path.getctime)

        # Copy the file to the destination directory
        shutil.copyfile(newest_xlsx_file, path)
        print(f"Newest XLSX file '{newest_xlsx_file}' copied to the directory.")
    # path = "Jul_2023.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Clear Original Filters
    if ws.auto_filter:
        ws.auto_filter.ref = None
        for i in range(2, ws.max_row + 1):
            ws.row_dimensions[i].hidden = False

    # Remove Original Colors and Border
    white_fill = PatternFill(fill_type='solid', start_color='FFFFFF')
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    for row in ws.rows:
        for cell in row:
            cell.border = thin_border
            cell.fill = white_fill


    # Add Daily Updates
    data = pd.read_csv("signal/" + time.strftime("%b_%Y") + '/' + time.strftime("%m_%d_%Y") + "_signal.csv")
    # data = pd.read_csv("07_29_2023_signal.csv")
    # data = pd.read_csv('full_result.csv')

    for wsrow in ws.iter_rows(min_row=2):
        for dfindex,dfrow in data.iterrows():
            if wsrow[0].value == dfrow['Stock']:
                wsrow[2].value = dfrow['Number of Trades']
                wsrow[3].value = dfrow['Profit']
                wsrow[4].value = dfrow['Profit %']
                wsrow[5].value = dfrow['Profit Factor']
                if (dfrow['BUY'] == 1):
                    wsrow[6].value = prev_time.strftime("%d/%m/%y")
                    wsrow[6].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    for i in range(5):
                        wsrow[i].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                if (dfrow['SELL'] == 1):
                    wsrow[7].value = prev_time.strftime("%d/%m/%y")   
                    wsrow[7].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') 
                    for i in range(5):
                        wsrow[i].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                if (dfrow['CLOSE BUY'] == 1):
                    wsrow[8].value = prev_time.strftime("%d/%m/%y")
                    wsrow[8].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    for i in range(5):
                        wsrow[i].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                if (dfrow['CLOSE SELL'] == 1):
                    wsrow[9].value = prev_time.strftime("%d/%m/%y")
                    wsrow[9].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    for i in range(5):
                        wsrow[i].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                wsrow[10].value = dfrow['Last Close']

    wb.save(path)


def trade_history():
    time = datetime.now()
    prev_time = time - timedelta(days=1)
    path = 'History.xlsx'
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    data = pd.read_csv("signal/" + time.strftime("%b_%Y") + '/' + time.strftime("%m_%d_%Y") + "_signal.csv")
    # data = pd.read_csv("signal/Jan_2024/01_03_2024_signal.csv")
    date = prev_time.strftime("%Y/%m/%d")

    for index, row in data.iterrows():

        # Append new trades
        if row['BUY'] == True or row['SELL'] == True:
            new_data = [row['Stock'], row['Number of Trades'], row['Profit'], row['Profit %'], row['Profit Factor']]
            if row['BUY'] == True:
                signal = [date, '']
            elif row['SELL'] == True:
                signal = ['', date]
            new_data.extend(signal)
            ws.append(new_data)
            
        # Close positions
        # Housekeep after closing position    
        if row['CLOSE BUY'] == True or row['CLOSE SELL'] == True:
            if row['CLOSE BUY'] == True:
                for index, wsrow in enumerate(ws.iter_rows(min_row=0)):
                    if wsrow[0].value == row['Stock'] and ws.row_dimensions[index+1].hidden == False:
                        wsrow[7].value = date
                        ws.row_dimensions[index+1].hidden = True
                        
            elif row['CLOSE SELL'] == True:
                for index, wsrow in enumerate(ws.iter_rows(min_row=0)):
                    if wsrow[0].value == row['Stock'] and ws.row_dimensions[index+1].hidden == False:
                        wsrow[8].value = date
                        ws.row_dimensions[index+1].hidden = True

    # Remove border
    no_border = Border(left=Side(), 
                        right=Side(), 
                        top=Side(), 
                        bottom=Side())

    for row in ws.rows:
        for cell in row:
            cell.border = no_border

    # Center cells
    for row in ws.iter_rows():
        for col in range(1,20):
            row[col].alignment = Alignment(horizontal='center')


    wb.save(path)


def signal_public():
    time = datetime.now()
    prev_time = time - timedelta(days=1)
    path = time.strftime("%b_%Y") + "_signal.xlsx"
    if not os.path.exists(path):
        # Get the newest XLSX file
        newest_xlsx_file = max(glob.glob("*.xlsx"), key=os.path.getctime)
        # Copy the file to the destination directory
        shutil.copyfile(newest_xlsx_file, path)
        print(f"Newest XLSX file '{newest_xlsx_file}' copied to the directory.")
    # path = "Jul_2023.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Remove Original Colors
    white_fill = PatternFill(fill_type='solid', start_color='FFFFFF')
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    for row in ws.rows:
        for cell in row:
            cell.border = thin_border
            cell.fill = white_fill


    # Add Daily Updates
    data = pd.read_csv("signal/" + time.strftime("%b_%Y") + '/' + time.strftime("%m_%d_%Y") + "_signal.csv")
    # data = pd.read_csv("signal/Dec_2023/12_01_2023_signal.csv")
    # data = pd.read_csv('full_result.csv')

    for wsrow in ws.iter_rows(min_row=2):
        for dfindex,dfrow in data.iterrows():
            if wsrow[0].value == dfrow['Stock']:
                wsrow[2].value = dfrow['Number of Trades']
                wsrow[3].value = dfrow['Profit']
                wsrow[4].value = dfrow['Profit %']
                wsrow[5].value = dfrow['Profit Factor']
                if (dfrow['BUY'] == 1):
                    wsrow[6].value = prev_time.strftime("%d/%m/%y")
                    wsrow[6].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    for i in range(5):
                        wsrow[i].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                if (dfrow['SELL'] == 1):
                    wsrow[7].value = prev_time.strftime("%d/%m/%y")   
                    wsrow[7].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid') 
                    for i in range(5):
                        wsrow[i].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                if (dfrow['CLOSE BUY'] == 1):
                    wsrow[8].value = prev_time.strftime("%d/%m/%y")
                    wsrow[8].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    for i in range(5):
                        wsrow[i].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                if (dfrow['CLOSE SELL'] == 1):
                    wsrow[9].value = prev_time.strftime("%d/%m/%y")
                    wsrow[9].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    for i in range(5):
                        wsrow[i].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    wb.save(path)




# signal()
# trade_history()
# signal_public()